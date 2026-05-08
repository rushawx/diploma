import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Define frontend options and their mapped backend API endpoints
frontend_options = [
    {
        "frontend_option": "Welcome Page",
        "description": "Static landing page with app information",
        "auth_required": False,
        "backend_endpoints": "None",
        "notes": "No backend calls required"
    },
    {
        "frontend_option": "Find Projects by Query",
        "description": "Search projects using transformer-based semantic search with a query string",
        "auth_required": True,
        "backend_endpoints": "GET /profile/me, GET /projects/ratings/all, PUT /projects/{project_id}",
        "notes": "Uses local transformer model for search; filters out already-rated projects"
    },
    {
        "frontend_option": "Find Projects by Tags",
        "description": "Search projects by selecting multiple tags using cosine similarity",
        "auth_required": True,
        "backend_endpoints": "GET /profile/me, GET /projects/ratings/all, GET /projects/with-tags, PUT /projects/{project_id}",
        "notes": "Local tag vector compilation; filters out already-rated projects"
    },
    {
        "frontend_option": "Get Recommendations",
        "description": "Get personalized project recommendations based on avatar's preferences using LightFM/ALS models",
        "auth_required": True,
        "backend_endpoints": "GET /profile/me, GET /avatars/my, GET /avatars/recommend, PUT /avatars/select/{avatar_id}, PUT /projects/{project_id}",
        "notes": "Uses local LightFM model; requires avatar to be associated first"
    },
    {
        "frontend_option": "View Available Projects",
        "description": "View list of all available projects",
        "auth_required": True,
        "backend_endpoints": "GET /projects/",
        "notes": "Displays all non-deleted projects"
    },
    {
        "frontend_option": "View Project Details",
        "description": "View detailed information about a specific project",
        "auth_required": True,
        "backend_endpoints": "GET /projects/, GET /projects/{project_id}",
        "notes": "User selects project from dropdown then fetches full details"
    },
    {
        "frontend_option": "View My Projects",
        "description": "View projects claimed by the current user with rating capability",
        "auth_required": True,
        "backend_endpoints": "GET /profile/me, GET /projects/user/{user_id}, GET /projects/{project_id}/rating, POST /projects/rate",
        "notes": "Shows user's claimed projects with interactive rating (1-5 stars)"
    },
    {
        "frontend_option": "Create Project",
        "description": "Create a new project with automatic embedding generation",
        "auth_required": True,
        "backend_endpoints": "POST /projects/",
        "notes": "Uses local transformer model to generate embeddings from project text"
    },
    {
        "frontend_option": "Your Avatar",
        "description": "View, select, or change avatar based on tag-based recommendations",
        "auth_required": True,
        "backend_endpoints": "GET /profile/me, GET /avatars/my, GET /tags/user/{user_id}, DELETE /tags/user/{user_id}, POST /tags/user/{user_id}, GET /avatars/recommend, GET /avatars/, PUT /avatars/select/{avatar_id}",
        "notes": "Allows updating user tags, viewing recommended avatars, and associating with avatar"
    },
    {
        "frontend_option": "Signup",
        "description": "Create new user account with profile data, tags, and auto-assign avatar",
        "auth_required": False,
        "backend_endpoints": "POST /profile/signup, POST /profile/login, GET /profile/me, POST /tags/user/{user_id}, GET /avatars/recommend, PUT /avatars/select/{avatar_id}",
        "notes": "After signup, automatically logs in, assigns tags, and assigns best-matching avatar"
    },
    {
        "frontend_option": "Login",
        "description": "Authenticate user with username and password",
        "auth_required": False,
        "backend_endpoints": "POST /profile/login",
        "notes": "Stores JWT token in session state for subsequent authenticated requests"
    },
    {
        "frontend_option": "Profile",
        "description": "View current user's profile information and associated avatar",
        "auth_required": True,
        "backend_endpoints": "GET /profile/me, GET /avatars/my",
        "notes": "Displays user details and avatar information"
    },
]

# Create workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Frontend Options"

# Set column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 70
ws.column_dimensions['E'].width = 40

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Write headers
headers = ["Frontend Option", "Description", "Auth Required", "Backend API Endpoints Used", "Notes"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border_style

# Write data
for row_idx, option in enumerate(frontend_options, 2):
    # Option name
    ws.cell(row=row_idx, column=1, value=option["frontend_option"]).alignment = cell_alignment
    ws.cell(row=row_idx, column=1).border = border_style

    # Description
    ws.cell(row=row_idx, column=2, value=option["description"]).alignment = cell_alignment
    ws.cell(row=row_idx, column=2).border = border_style

    # Auth required
    ws.cell(row=row_idx, column=3, value=option["auth_required"]).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_idx, column=3).border = border_style

    # Backend endpoints
    ws.cell(row=row_idx, column=4, value=option["backend_endpoints"]).alignment = cell_alignment
    ws.cell(row=row_idx, column=4).border = border_style

    # Notes
    ws.cell(row=row_idx, column=5, value=option["notes"]).alignment = cell_alignment
    ws.cell(row=row_idx, column=5).border = border_style

# Apply row height for better readability
for row_idx in range(2, len(frontend_options) + 2):
    ws.row_dimensions[row_idx].height = 60

# Save workbook
output_path = "/Users/antonshishkov/Projects/diploma/frontend_options.xlsx"
wb.save(output_path)
print(f"Excel file created successfully at: {output_path}")
print(f"Total frontend options: {len(frontend_options)}")

# Summary stats
auth_required = sum(1 for o in frontend_options if o["auth_required"])
print(f"Options requiring authentication: {auth_required}")
print(f"Options without authentication: {len(frontend_options) - auth_required}")
