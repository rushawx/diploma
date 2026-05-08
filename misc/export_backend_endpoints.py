import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Define all endpoints with their descriptions
endpoints = [
    # Auth endpoints
    {
        "method": "GET",
        "path": "/",
        "description": "Root endpoint, returns API information",
        "auth_required": False,
        "module": "main.py"
    },
    {
        "method": "POST",
        "path": "/token",
        "description": "Login endpoint using OAuth2PasswordRequestForm, returns JWT bearer token",
        "auth_required": False,
        "module": "main.py"
    },
    {
        "method": "GET",
        "path": "/about_user",
        "description": "Get info about current authenticated user",
        "auth_required": True,
        "module": "main.py"
    },
    # Profile endpoints
    {
        "method": "POST",
        "path": "/profile/signup",
        "description": "Create a new user account with profile information",
        "auth_required": False,
        "module": "profile.py"
    },
    {
        "method": "POST",
        "path": "/profile/login",
        "description": "User login endpoint, returns JWT token (avatars cannot login)",
        "auth_required": False,
        "module": "profile.py"
    },
    {
        "method": "GET",
        "path": "/profile/me",
        "description": "Get current authenticated user's profile information",
        "auth_required": True,
        "module": "profile.py"
    },
    # Projects endpoints
    {
        "method": "GET",
        "path": "/projects/",
        "description": "Get all projects (non-deleted)",
        "auth_required": True,
        "module": "projects.py"
    },
    {
        "method": "GET",
        "path": "/projects/with-embeddings",
        "description": "Get all projects with embeddings for similarity search",
        "auth_required": True,
        "module": "projects.py"
    },
    {
        "method": "GET",
        "path": "/projects/with-tags",
        "description": "Get all projects with tags for tag-based similarity search",
        "auth_required": True,
        "module": "projects.py"
    },
    {
        "method": "GET",
        "path": "/projects/user/{user_id}",
        "description": "Get all projects chosen by a specific user",
        "auth_required": True,
        "module": "projects.py"
    },
    {
        "method": "GET",
        "path": "/projects/{project_id}",
        "description": "Get a specific project by ID",
        "auth_required": True,
        "module": "projects.py"
    },
    {
        "method": "POST",
        "path": "/projects/",
        "description": "Create a new project",
        "auth_required": True,
        "module": "projects.py"
    },
    {
        "method": "PUT",
        "path": "/projects/{project_id}",
        "description": "Update an existing project",
        "auth_required": True,
        "module": "projects.py"
    },
    {
        "method": "DELETE",
        "path": "/projects/{project_id}",
        "description": "Soft delete a project (sets deleted_at timestamp)",
        "auth_required": True,
        "module": "projects.py"
    },
    {
        "method": "POST",
        "path": "/projects/rate",
        "description": "Rate a project (1-5 stars), updates existing rating if present",
        "auth_required": True,
        "module": "projects.py"
    },
    {
        "method": "GET",
        "path": "/projects/{project_id}/rating",
        "description": "Get current user's rating for a specific project",
        "auth_required": True,
        "module": "projects.py"
    },
    {
        "method": "GET",
        "path": "/projects/{project_id}/ratings",
        "description": "Get all ratings for a specific project",
        "auth_required": True,
        "module": "projects.py"
    },
    {
        "method": "GET",
        "path": "/projects/ratings/all",
        "description": "Get all ratings in format suitable for ALS model",
        "auth_required": True,
        "module": "projects.py"
    },
    # Users endpoints
    {
        "method": "GET",
        "path": "/users/",
        "description": "Get all users, optionally filtered by user_type (student/avatar)",
        "auth_required": True,
        "module": "users.py"
    },
    {
        "method": "GET",
        "path": "/users/avatars",
        "description": "Get all avatar-type users",
        "auth_required": True,
        "module": "users.py"
    },
    {
        "method": "GET",
        "path": "/users/{user_id}",
        "description": "Get a specific user by ID",
        "auth_required": True,
        "module": "users.py"
    },
    {
        "method": "POST",
        "path": "/users/",
        "description": "Create a new user (NOT IMPLEMENTED)",
        "auth_required": True,
        "module": "users.py"
    },
    {
        "method": "PUT",
        "path": "/users/{user_id}",
        "description": "Update a user (NOT IMPLEMENTED)",
        "auth_required": True,
        "module": "users.py"
    },
    {
        "method": "DELETE",
        "path": "/users/{user_id}",
        "description": "Delete a user (NOT IMPLEMENTED)",
        "auth_required": True,
        "module": "users.py"
    },
    # Avatars endpoints
    {
        "method": "GET",
        "path": "/avatars/my",
        "description": "Get current user's associated avatar",
        "auth_required": True,
        "module": "avatars.py"
    },
    {
        "method": "GET",
        "path": "/avatars/recommend",
        "description": "Recommend avatars based on user's tags similarity (cosine similarity)",
        "auth_required": True,
        "module": "avatars.py"
    },
    {
        "method": "GET",
        "path": "/avatars/",
        "description": "Get all avatar-type users",
        "auth_required": True,
        "module": "avatars.py"
    },
    {
        "method": "GET",
        "path": "/avatars/{avatar_id}",
        "description": "Get a specific avatar by ID",
        "auth_required": True,
        "module": "avatars.py"
    },
    {
        "method": "PUT",
        "path": "/avatars/select/{avatar_id}",
        "description": "Associate current user with an avatar",
        "auth_required": True,
        "module": "avatars.py"
    },
    # Tags endpoints
    {
        "method": "GET",
        "path": "/tags/",
        "description": "Get all tags with pagination (skip/limit)",
        "auth_required": True,
        "module": "tags.py"
    },
    {
        "method": "GET",
        "path": "/tags/{tag_id}",
        "description": "Get a specific tag by ID",
        "auth_required": True,
        "module": "tags.py"
    },
    {
        "method": "GET",
        "path": "/tags/user/{user_id}",
        "description": "Get all tags for a specific user",
        "auth_required": True,
        "module": "tags.py"
    },
    {
        "method": "GET",
        "path": "/tags/project/{project_id}",
        "description": "Get all tags for a specific project",
        "auth_required": True,
        "module": "tags.py"
    },
    {
        "method": "POST",
        "path": "/tags/user/{user_id}",
        "description": "Add multiple tags to a user by tag names",
        "auth_required": True,
        "module": "tags.py"
    },
    {
        "method": "DELETE",
        "path": "/tags/user/{user_id}",
        "description": "Delete all tags for a user",
        "auth_required": True,
        "module": "tags.py"
    },
]

# Create workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "API Endpoints"

# Set column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 60
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 20

# Define header style
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Define cell alignment
cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Write headers
headers = ["Method", "Path", "Description", "Auth Required", "Module"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Write data
for row_idx, endpoint in enumerate(endpoints, 2):
    ws.cell(row=row_idx, column=1, value=endpoint["method"]).alignment = Alignment(horizontal="center")
    ws.cell(row=row_idx, column=2, value=endpoint["path"]).alignment = cell_alignment
    ws.cell(row=row_idx, column=3, value=endpoint["description"]).alignment = cell_alignment
    ws.cell(row=row_idx, column=4, value=endpoint["auth_required"]).alignment = Alignment(horizontal="center")
    ws.cell(row=row_idx, column=5, value=endpoint["module"]).alignment = cell_alignment

# Apply method color coding
for row_idx, endpoint in enumerate(endpoints, 2):
    method_cell = ws.cell(row=row_idx, column=1)
    if endpoint["method"] == "GET":
        method_cell.fill = PatternFill(start_color="E7F6FF", end_color="E7F6FF", fill_type="solid")
    elif endpoint["method"] == "POST":
        method_cell.fill = PatternFill(start_color="E2FBD7", end_color="E2FBD7", fill_type="solid")
    elif endpoint["method"] == "PUT":
        method_cell.fill = PatternFill(start_color="FFF4CE", end_color="FFF4CE", fill_type="solid")
    elif endpoint["method"] == "DELETE":
        method_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

# Save workbook
output_path = "api_backend_endpoints.xlsx"
wb.save(output_path)
print(f"Excel file created successfully at: {output_path}")
print(f"Total endpoints: {len(endpoints)}")
